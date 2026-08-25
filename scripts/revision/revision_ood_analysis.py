"""Leakage-free scaffold-held-out baseline and chemical-support analyses.

The script deliberately rebuilds every input from the source XLSX and writes
only CSV/NPZ/JSON outputs.  It does not deserialize the project's historical
pickle checkpoints.

Analyses
--------
1. Two repeats of five-fold stratified scaffold-group cross-validation.
2. Hyperparameter choice by scaffold-grouped inner CV (outer test untouched).
3. ROC-AUC, PR-AUC, balanced accuracy, sensitivity, specificity, Brier score,
   calibration slope/intercept, and expected calibration error.
4. Test-to-training nearest-neighbour Tanimoto similarity and performance
   trends across similarity strata.
5. Scaffold-cluster bootstrap confidence intervals from pooled out-of-fold
   predictions.

The primary representation is the manuscript's 2,048-bit radius-2 Morgan
fingerprint.  The split construction treats a molecule with an empty Murcko
scaffold as its own canonical-SMILES framework rather than placing all acyclic
molecules into one artificial mega-group.
"""

from __future__ import annotations

import argparse
import json
import math
import sys
import time
from dataclasses import dataclass
from pathlib import Path

import lightgbm as lgb
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from rdkit import Chem, DataStructs, RDLogger
from rdkit.Chem import rdFingerprintGenerator
from rdkit.Chem.Scaffolds import MurckoScaffold
from scipy import stats
from scipy.sparse import csr_matrix, load_npz, save_npz
from sklearn.calibration import calibration_curve
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    average_precision_score,
    balanced_accuracy_score,
    brier_score_loss,
    confusion_matrix,
    roc_auc_score,
)
from sklearn.model_selection import StratifiedGroupKFold, train_test_split


RDLogger.DisableLog("rdApp.*")

ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "data" / "toxnew_right.xlsx"
OUT = ROOT / "results" / "revision" / "ood_analysis"
FEATURES = OUT / "morgan_features.npz"
RECORDS = OUT / "records.csv"

ENDPOINT_NAMES = {
    "prenatal_development_C": "Prenatal developmental toxicity",
    "TSHR_agonist_activity_C": "TSHR agonist",
    "respiratory_toxicity_C": "Respiratory toxicity",
    "ocular_toxicity_C": "Ocular toxicity",
    "ames_mutagenicity_C": "Ames mutagenicity",
    "reproductive_toxicity_C": "Reproductive toxicity",
    "skin_corrosion_C": "Skin corrosion",
    "neurotoxicity_C": "Neurotoxicity",
    "Estrogen_Receptor_α_C": "Estrogen receptor α",
    "Androgen_Receptor_C": "Androgen receptor",
    "cytotoxicity_C": "Cytotoxicity",
    "Carcinogenicity_C": "Carcinogenicity",
    "Hepatotoxicity_C": "Hepatotoxicity",
}

# Small, prespecified candidate set.  Every candidate is compared only within
# each outer-training set by scaffold-grouped inner CV.
PARAM_CANDIDATES = [
    dict(n_estimators=250, max_depth=5, num_leaves=31, learning_rate=0.05,
         min_child_samples=20, subsample=0.8, colsample_bytree=0.8,
         reg_alpha=0.0, reg_lambda=0.0),
    dict(n_estimators=500, max_depth=7, num_leaves=63, learning_rate=0.03,
         min_child_samples=20, subsample=0.8, colsample_bytree=0.8,
         reg_alpha=0.0, reg_lambda=0.1),
    dict(n_estimators=350, max_depth=9, num_leaves=127, learning_rate=0.05,
         min_child_samples=10, subsample=0.9, colsample_bytree=0.7,
         reg_alpha=0.01, reg_lambda=0.1),
    dict(n_estimators=700, max_depth=5, num_leaves=31, learning_rate=0.02,
         min_child_samples=30, subsample=0.9, colsample_bytree=0.9,
         reg_alpha=0.1, reg_lambda=1.0),
    dict(n_estimators=300, max_depth=-1, num_leaves=63, learning_rate=0.05,
         min_child_samples=30, subsample=0.8, colsample_bytree=0.8,
         reg_alpha=0.1, reg_lambda=0.1),
    dict(n_estimators=500, max_depth=11, num_leaves=127, learning_rate=0.03,
         min_child_samples=50, subsample=0.7, colsample_bytree=0.9,
         reg_alpha=1.0, reg_lambda=0.1),
]


def log(message: str) -> None:
    print(message, flush=True)


def canonicalize(smiles: str) -> tuple[str, Chem.Mol | None]:
    mol = Chem.MolFromSmiles(str(smiles))
    if mol is None:
        return "", None
    return Chem.MolToSmiles(mol, canonical=True), mol


def scaffold_for(mol: Chem.Mol, canonical_smiles: str) -> str:
    scaffold = MurckoScaffold.MurckoScaffoldSmiles(
        mol=mol, includeChirality=False
    )
    return scaffold if scaffold else f"ACYCLIC::{canonical_smiles}"


def build_features() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    log(f"Reading {SOURCE}")
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    fpgen = rdFingerprintGenerator.GetMorganGenerator(radius=2, fpSize=2048)
    rows: list[dict[str, object]] = []
    sparse_rows: list[np.ndarray] = []
    invalid: list[dict[str, object]] = []

    for endpoint in workbook.sheetnames:
        sheet = workbook[endpoint]
        for source_row, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not values or values[0] is None or values[1] is None:
                continue
            canonical, mol = canonicalize(str(values[0]))
            if mol is None:
                invalid.append(
                    {"endpoint": endpoint, "source_row": source_row, "smiles": values[0]}
                )
                continue
            label = int(values[1])
            if label not in (0, 1):
                invalid.append(
                    {"endpoint": endpoint, "source_row": source_row, "smiles": values[0]}
                )
                continue
            fp = fpgen.GetFingerprint(mol)
            onbits = np.asarray(list(fp.GetOnBits()), dtype=np.int32)
            sparse_rows.append(onbits)
            rows.append(
                {
                    "record_id": len(rows),
                    "endpoint": endpoint,
                    "endpoint_name": ENDPOINT_NAMES.get(endpoint, endpoint),
                    "source_row": source_row,
                    "smiles": canonical,
                    "label": label,
                    "scaffold": scaffold_for(mol, canonical),
                }
            )
        log(f"  {endpoint}: {sum(r['endpoint'] == endpoint for r in rows)} valid records")

    indptr = [0]
    indices: list[int] = []
    for onbits in sparse_rows:
        indices.extend(onbits.tolist())
        indptr.append(len(indices))
    matrix = csr_matrix(
        (
            np.ones(len(indices), dtype=np.uint8),
            np.asarray(indices, dtype=np.int32),
            np.asarray(indptr, dtype=np.int64),
        ),
        shape=(len(rows), 2048),
        dtype=np.uint8,
    )
    frame = pd.DataFrame(rows)
    frame.to_csv(RECORDS, index=False, encoding="utf-8-sig")
    save_npz(FEATURES, matrix, compressed=True)
    pd.DataFrame(invalid).to_csv(OUT / "invalid_records.csv", index=False, encoding="utf-8-sig")
    summary = {
        "records": int(len(frame)),
        "endpoints": int(frame.endpoint.nunique()),
        "invalid_records": int(len(invalid)),
        "unique_canonical_smiles": int(frame.smiles.nunique()),
        "unique_scaffolds": int(frame.scaffold.nunique()),
        "feature_shape": list(matrix.shape),
    }
    (OUT / "feature_summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    log(f"Saved {matrix.shape} Morgan matrix and {len(frame)} records")


def make_model(params: dict[str, object], seed: int) -> lgb.LGBMClassifier:
    return lgb.LGBMClassifier(
        **params,
        objective="binary",
        random_state=seed,
        n_jobs=8,
        verbosity=-1,
        deterministic=True,
        force_col_wise=True,
    )


def choose_params(
    x: csr_matrix,
    y: np.ndarray,
    groups: np.ndarray,
    seed: int,
) -> tuple[int, list[float]]:
    splitter = StratifiedGroupKFold(n_splits=3, shuffle=True, random_state=seed)
    candidate_scores: list[float] = []
    for candidate_id, params in enumerate(PARAM_CANDIDATES):
        scores: list[float] = []
        for inner_train, inner_valid in splitter.split(x, y, groups):
            if np.unique(y[inner_valid]).size < 2:
                continue
            model = make_model(params, seed + candidate_id)
            model.fit(x[inner_train], y[inner_train])
            probability = model.predict_proba(x[inner_valid])[:, 1]
            scores.append(roc_auc_score(y[inner_valid], probability))
        candidate_scores.append(float(np.mean(scores)) if scores else 0.5)
    return int(np.argmax(candidate_scores)), candidate_scores


def fp_from_sparse_row(x: csr_matrix, row: int) -> DataStructs.ExplicitBitVect:
    fp = DataStructs.ExplicitBitVect(x.shape[1])
    for bit in x.indices[x.indptr[row] : x.indptr[row + 1]]:
        fp.SetBit(int(bit))
    return fp


def nearest_tanimoto(
    x: csr_matrix, train_indices: np.ndarray, test_indices: np.ndarray
) -> np.ndarray:
    train_fps = [fp_from_sparse_row(x, int(i)) for i in train_indices]
    values = np.empty(len(test_indices), dtype=np.float32)
    for output_index, row_index in enumerate(test_indices):
        test_fp = fp_from_sparse_row(x, int(row_index))
        similarities = DataStructs.BulkTanimotoSimilarity(test_fp, train_fps)
        values[output_index] = max(similarities) if similarities else np.nan
    return values


def safe_auc(y: np.ndarray, p: np.ndarray) -> float:
    return float(roc_auc_score(y, p)) if np.unique(y).size == 2 else math.nan


def calibration_stats(y: np.ndarray, p: np.ndarray) -> tuple[float, float, float]:
    clipped = np.clip(p, 1e-6, 1 - 1e-6)
    logit = np.log(clipped / (1 - clipped)).reshape(-1, 1)
    if np.unique(y).size < 2:
        return math.nan, math.nan, math.nan
    calibrator = LogisticRegression(C=1e6, solver="lbfgs")
    calibrator.fit(logit, y)
    intercept = float(calibrator.intercept_[0])
    slope = float(calibrator.coef_[0, 0])
    # Equal-width ECE.  Empty bins contribute zero weight.
    edges = np.linspace(0, 1, 11)
    ece = 0.0
    for lower, upper in zip(edges[:-1], edges[1:]):
        mask = (p >= lower) & (p < upper if upper < 1 else p <= upper)
        if mask.any():
            ece += mask.mean() * abs(float(y[mask].mean()) - float(p[mask].mean()))
    return intercept, slope, float(ece)


def metrics(y: np.ndarray, p: np.ndarray) -> dict[str, float]:
    predicted = (p >= 0.5).astype(int)
    tn, fp, fn, tp = confusion_matrix(y, predicted, labels=[0, 1]).ravel()
    intercept, slope, ece = calibration_stats(y, p)
    return {
        "roc_auc": safe_auc(y, p),
        "pr_auc": float(average_precision_score(y, p)),
        "balanced_accuracy": float(balanced_accuracy_score(y, predicted)),
        "sensitivity": float(tp / (tp + fn)) if tp + fn else math.nan,
        "specificity": float(tn / (tn + fp)) if tn + fp else math.nan,
        "brier_score": float(brier_score_loss(y, p)),
        "calibration_intercept": intercept,
        "calibration_slope": slope,
        "ece_10": ece,
    }


def random_split_similarity(
    x: csr_matrix, y: np.ndarray, seed: int
) -> np.ndarray:
    all_indices = np.arange(len(y))
    train_indices, test_indices = train_test_split(
        all_indices, test_size=0.2, stratify=y, random_state=seed
    )
    similarities = nearest_tanimoto(x, train_indices, test_indices)
    return similarities


def run_outer_cv(smoke: bool = False) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    records = pd.read_csv(RECORDS, encoding="utf-8-sig")
    matrix = load_npz(FEATURES).tocsr().astype(np.float32)
    prediction_rows: list[dict[str, object]] = []
    fold_rows: list[dict[str, object]] = []
    parameter_rows: list[dict[str, object]] = []
    similarity_rows: list[dict[str, object]] = []
    endpoint_values = records.endpoint.drop_duplicates().tolist()
    if smoke:
        endpoint_values = endpoint_values[:1]

    for endpoint_index, endpoint in enumerate(endpoint_values, start=1):
        endpoint_mask = records.endpoint.eq(endpoint).to_numpy()
        global_indices = np.flatnonzero(endpoint_mask)
        endpoint_frame = records.loc[endpoint_mask].reset_index(drop=True)
        x = matrix[global_indices]
        y = endpoint_frame.label.to_numpy(dtype=np.int8)
        groups = endpoint_frame.scaffold.to_numpy(dtype=str)
        log(f"[{endpoint_index}/{len(endpoint_values)}] {endpoint} n={len(y)} groups={len(set(groups))}")

        for repeat, seed in enumerate((101, 202), start=1):
            outer = StratifiedGroupKFold(n_splits=5, shuffle=True, random_state=seed)
            for fold, (train_indices, test_indices) in enumerate(
                outer.split(x, y, groups), start=1
            ):
                t0 = time.time()
                chosen, candidate_scores = choose_params(
                    x[train_indices], y[train_indices], groups[train_indices], seed + fold
                )
                model = make_model(PARAM_CANDIDATES[chosen], seed + fold)
                model.fit(x[train_indices], y[train_indices])
                probability = model.predict_proba(x[test_indices])[:, 1]
                similarities = nearest_tanimoto(x, train_indices, test_indices)
                fold_metric = metrics(y[test_indices], probability)
                fold_rows.append(
                    {
                        "endpoint": endpoint,
                        "endpoint_name": ENDPOINT_NAMES.get(endpoint, endpoint),
                        "repeat": repeat,
                        "fold": fold,
                        "n_train": len(train_indices),
                        "n_test": len(test_indices),
                        "positive_fraction_test": float(y[test_indices].mean()),
                        "n_train_scaffolds": int(np.unique(groups[train_indices]).size),
                        "n_test_scaffolds": int(np.unique(groups[test_indices]).size),
                        "scaffold_overlap": int(
                            len(set(groups[train_indices]).intersection(groups[test_indices]))
                        ),
                        "median_nn_tanimoto": float(np.median(similarities)),
                        "chosen_candidate": chosen,
                        "elapsed_seconds": time.time() - t0,
                        **fold_metric,
                    }
                )
                parameter_rows.append(
                    {
                        "endpoint": endpoint,
                        "repeat": repeat,
                        "fold": fold,
                        "chosen_candidate": chosen,
                        "candidate_scores": json.dumps(candidate_scores),
                        "parameters": json.dumps(PARAM_CANDIDATES[chosen]),
                    }
                )
                for local_index, p, similarity in zip(test_indices, probability, similarities):
                    row = endpoint_frame.iloc[int(local_index)]
                    prediction_rows.append(
                        {
                            "endpoint": endpoint,
                            "endpoint_name": ENDPOINT_NAMES.get(endpoint, endpoint),
                            "repeat": repeat,
                            "fold": fold,
                            "record_id": int(row.record_id),
                            "endpoint_row": int(local_index),
                            "smiles": row.smiles,
                            "scaffold": row.scaffold,
                            "label": int(row.label),
                            "probability": float(p),
                            "nn_tanimoto": float(similarity),
                        }
                    )
                log(
                    f"  repeat {repeat} fold {fold}: AUC={fold_metric['roc_auc']:.3f}, "
                    f"PR={fold_metric['pr_auc']:.3f}, medianNN={np.median(similarities):.3f}, "
                    f"candidate={chosen}, {time.time()-t0:.1f}s"
                )
                pd.DataFrame(prediction_rows).to_csv(
                    OUT / "independent_cv_predictions.partial.csv", index=False, encoding="utf-8-sig"
                )
                pd.DataFrame(fold_rows).to_csv(
                    OUT / "independent_cv_fold_metrics.partial.csv", index=False, encoding="utf-8-sig"
                )

            random_sim = random_split_similarity(x, y, seed)
            similarity_rows.extend(
                {
                    "endpoint": endpoint,
                    "endpoint_name": ENDPOINT_NAMES.get(endpoint, endpoint),
                    "repeat": repeat,
                    "protocol": "random 80/20",
                    "nn_tanimoto": float(value),
                }
                for value in random_sim
            )

        # Preserve the independently randomized repeat identifier in the
        # exported protocol-comparison source data. Earlier exports pooled
        # both repeats correctly for the distribution but replaced the repeat
        # field with NaN, which made row-level provenance harder to audit.
        endpoint_scaffold_rows = [
            row for row in prediction_rows if row["endpoint"] == endpoint
        ]
        similarity_rows.extend(
            {
                "endpoint": endpoint,
                "endpoint_name": ENDPOINT_NAMES.get(endpoint, endpoint),
                "repeat": int(row["repeat"]),
                "protocol": "repeated scaffold-group CV",
                "nn_tanimoto": float(row["nn_tanimoto"]),
            }
            for row in endpoint_scaffold_rows
        )

    predictions = pd.DataFrame(prediction_rows)
    fold_metrics = pd.DataFrame(fold_rows)
    predictions.to_csv(OUT / "independent_cv_predictions.csv", index=False, encoding="utf-8-sig")
    fold_metrics.to_csv(OUT / "independent_cv_fold_metrics.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(parameter_rows).to_csv(
        OUT / "nested_cv_parameters.csv", index=False, encoding="utf-8-sig"
    )
    pd.DataFrame(similarity_rows).to_csv(
        OUT / "similarity_protocol_comparison.csv", index=False, encoding="utf-8-sig"
    )
    summarize(predictions, fold_metrics, pd.DataFrame(similarity_rows), smoke=smoke)


def scaffold_bootstrap(
    frame: pd.DataFrame, n_bootstrap: int = 1000, seed: int = 814
) -> dict[str, tuple[float, float]]:
    rng = np.random.default_rng(seed)
    unique_scaffolds = frame.scaffold.unique()
    distributions: dict[str, list[float]] = {
        "roc_auc": [],
        "pr_auc": [],
        "balanced_accuracy": [],
        "sensitivity": [],
        "specificity": [],
        "brier_score": [],
    }
    grouped_indices = {
        scaffold: frame.index[frame.scaffold.eq(scaffold)].to_numpy()
        for scaffold in unique_scaffolds
    }
    for _ in range(n_bootstrap):
        sampled = rng.choice(unique_scaffolds, size=len(unique_scaffolds), replace=True)
        indices = np.concatenate([grouped_indices[value] for value in sampled])
        y = frame.loc[indices, "label"].to_numpy(dtype=int)
        p = frame.loc[indices, "probability"].to_numpy(dtype=float)
        if np.unique(y).size < 2:
            continue
        result = metrics(y, p)
        for key in distributions:
            distributions[key].append(result[key])
    return {
        key: (float(np.percentile(values, 2.5)), float(np.percentile(values, 97.5)))
        for key, values in distributions.items()
        if values
    }


def summarize(
    predictions: pd.DataFrame,
    fold_metrics: pd.DataFrame,
    similarity_protocols: pd.DataFrame,
    smoke: bool = False,
) -> None:
    # Average the two independently generated OOF probabilities for one record.
    pooled = (
        predictions.groupby(
            ["endpoint", "endpoint_name", "record_id", "smiles", "scaffold", "label"],
            as_index=False,
        )
        .agg(probability=("probability", "mean"), nn_tanimoto=("nn_tanimoto", "mean"))
    )
    pooled.to_csv(OUT / "pooled_oof_predictions.csv", index=False, encoding="utf-8-sig")

    endpoint_rows: list[dict[str, object]] = []
    bootstrap_rows: list[dict[str, object]] = []
    for endpoint, frame in pooled.groupby("endpoint", sort=False):
        values = metrics(frame.label.to_numpy(dtype=int), frame.probability.to_numpy(dtype=float))
        endpoint_rows.append(
            {
                "endpoint": endpoint,
                "endpoint_name": frame.endpoint_name.iloc[0],
                "n": len(frame),
                "positive_fraction": float(frame.label.mean()),
                "n_scaffolds": int(frame.scaffold.nunique()),
                "median_nn_tanimoto": float(frame.nn_tanimoto.median()),
                "q25_nn_tanimoto": float(frame.nn_tanimoto.quantile(0.25)),
                "q75_nn_tanimoto": float(frame.nn_tanimoto.quantile(0.75)),
                **values,
            }
        )
        intervals = scaffold_bootstrap(frame, n_bootstrap=100 if smoke else 1000)
        for metric_name, (lower, upper) in intervals.items():
            bootstrap_rows.append(
                {
                    "endpoint": endpoint,
                    "endpoint_name": frame.endpoint_name.iloc[0],
                    "metric": metric_name,
                    "estimate": values[metric_name],
                    "ci_lower": lower,
                    "ci_upper": upper,
                    "resampling_unit": "Bemis-Murcko scaffold group",
                    "n_bootstrap": 100 if smoke else 1000,
                }
            )

    endpoint_metrics = pd.DataFrame(endpoint_rows)
    endpoint_metrics.to_csv(OUT / "endpoint_metrics.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(bootstrap_rows).to_csv(
        OUT / "scaffold_bootstrap_ci.csv", index=False, encoding="utf-8-sig"
    )

    # Within each endpoint, quartiles provide comparable support across strata.
    trend_rows: list[dict[str, object]] = []
    for endpoint, frame in pooled.groupby("endpoint", sort=False):
        frame = frame.copy()
        frame["similarity_quartile"] = pd.qcut(
            frame.nn_tanimoto.rank(method="first"), 4, labels=["Q1", "Q2", "Q3", "Q4"]
        )
        for quartile, stratum in frame.groupby("similarity_quartile", observed=True):
            trend_rows.append(
                {
                    "endpoint": endpoint,
                    "endpoint_name": frame.endpoint_name.iloc[0],
                    "similarity_quartile": str(quartile),
                    "n": len(stratum),
                    "median_nn_tanimoto": float(stratum.nn_tanimoto.median()),
                    "roc_auc": safe_auc(
                        stratum.label.to_numpy(dtype=int),
                        stratum.probability.to_numpy(dtype=float),
                    ),
                    "pr_auc": float(
                        average_precision_score(stratum.label, stratum.probability)
                    ),
                    "brier_score": float(brier_score_loss(stratum.label, stratum.probability)),
                }
            )
    trend = pd.DataFrame(trend_rows)
    trend.to_csv(OUT / "performance_by_similarity_quartile.csv", index=False, encoding="utf-8-sig")
    valid_trend = trend.dropna(subset=["roc_auc"])
    rho_auc, p_auc = stats.spearmanr(valid_trend.median_nn_tanimoto, valid_trend.roc_auc)
    rho_brier, p_brier = stats.spearmanr(
        valid_trend.median_nn_tanimoto, valid_trend.brier_score
    )

    protocol_summary = (
        similarity_protocols.groupby("protocol").nn_tanimoto
        .agg(["count", "mean", "median", "std", lambda value: value.quantile(0.25), lambda value: value.quantile(0.75)])
        .reset_index()
    )
    protocol_summary.columns = ["protocol", "n", "mean", "median", "sd", "q25", "q75"]
    protocol_summary.to_csv(
        OUT / "similarity_protocol_summary.csv", index=False, encoding="utf-8-sig"
    )

    aggregate = {
        "n_endpoints": int(endpoint_metrics.endpoint.nunique()),
        "macro_mean_metrics": {
            column: float(endpoint_metrics[column].mean())
            for column in [
                "roc_auc",
                "pr_auc",
                "balanced_accuracy",
                "sensitivity",
                "specificity",
                "brier_score",
                "calibration_intercept",
                "calibration_slope",
                "ece_10",
            ]
        },
        "similarity_performance_trend": {
            "spearman_rho_similarity_vs_auc": float(rho_auc),
            "p_similarity_vs_auc": float(p_auc),
            "spearman_rho_similarity_vs_brier": float(rho_brier),
            "p_similarity_vs_brier": float(p_brier),
        },
        "fold_scaffold_overlap_max": int(fold_metrics.scaffold_overlap.max()),
    }
    (OUT / "aggregate_summary.json").write_text(
        json.dumps(aggregate, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    log(json.dumps(aggregate, indent=2, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--build-features", action="store_true")
    parser.add_argument("--run", action="store_true")
    parser.add_argument("--smoke", action="store_true")
    arguments = parser.parse_args()
    if arguments.build_features or not FEATURES.exists() or not RECORDS.exists():
        build_features()
    if arguments.run:
        run_outer_cv(smoke=arguments.smoke)


if __name__ == "__main__":
    main()
